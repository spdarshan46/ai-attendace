# ==========================================
#  ADMIN_PANEL.PY — Enterprise Admin
#  FIXES in this version:
#    ✅ Delete bug fixed — confirm dialog
#       shown BEFORE window is destroyed
#    ✅ All +22/+33/+44 alpha hacks removed
#    ✅ Light mode works via (light, dark) tuples
#    ✅ Settings theme switch rebuilds panel colors
# ==========================================

import customtkinter as ctk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timedelta
import csv, os

from ui.theme import *
from ui.components.widgets import StatCard, NeonButton
from database import employees, attendance, logs, write_log

FACES_DIR  = "faces"
LATE_AFTER = "09:30:00"


class AdminPanel(ctk.CTkFrame):

    def __init__(self, parent, on_back):
        super().__init__(parent, fg_color=BG_BASE, corner_radius=0)
        self.pack(fill="both", expand=True)
        self._on_back = on_back
        self._active  = None

        self._build_sidebar()
        self._build_topbar()
        self._build_body()
        self._select("dashboard")

    # ──────────────────────────────────────────────────────────────────────────
    #  SIDEBAR
    # ──────────────────────────────────────────────────────────────────────────
    def _build_sidebar(self):
        self._sidebar = ctk.CTkFrame(self, width=250,
                                      fg_color=SIDEBAR_BG, corner_radius=0)
        self._sidebar.pack(side="left", fill="y")
        self._sidebar.pack_propagate(False)

        # Logo
        logo = ctk.CTkFrame(self._sidebar, height=110, fg_color="transparent")
        logo.pack(fill="x")
        ctk.CTkLabel(logo, text="⬡", font=("Segoe UI", 34),
                     text_color=NEON_BLUE).pack(pady=(20, 0))
        ctk.CTkLabel(logo, text="ADMIN HUB",
                     font=("Courier New", 13, "bold"),
                     text_color=NEON_BLUE).pack()
        ctk.CTkLabel(logo, text="Enterprise Dashboard",
                     font=FONT_MICRO, text_color=TEXT_MUTED).pack()

        # Separator — solid dim color (no alpha)
        ctk.CTkFrame(self._sidebar, height=1,
                     fg_color=DIVIDER_COLOR).pack(fill="x", padx=PAD_MD)

        # Menu buttons
        self._menu_btns = {}
        items = [
            ("dashboard",  "⬡", "Dashboard"),
            ("employees",  "◈", "Employees"),
            ("attendance", "◷", "Attendance"),
            ("analytics",  "◉", "Analytics"),
            ("logs",       "◌", "System Logs"),
            ("settings",   "⚙", "Settings"),
        ]
        for key, icon, label in items:
            btn = ctk.CTkButton(
                self._sidebar,
                text=f"  {icon}   {label}",
                anchor="w", height=46,
                fg_color="transparent",
                hover_color=SIDEBAR_HOVER,
                corner_radius=RADIUS_MD,
                font=("Courier New", 13),
                text_color=TEXT_SECONDARY,
                command=lambda k=key: self._select(k)
            )
            btn.pack(fill="x", padx=PAD_SM, pady=2)
            self._menu_btns[key] = btn

        # Back — solid hover color (no alpha)
        ctk.CTkButton(
            self._sidebar, text="  ← Back to Home",
            anchor="w", height=44,
            fg_color="transparent",
            hover_color=NEON_PINK_MID,    # was DANGER + "33" — now solid
            corner_radius=RADIUS_MD,
            font=("Courier New", 12),
            text_color=NEON_PINK,
            command=self._on_back
        ).pack(side="bottom", fill="x", padx=PAD_SM, pady=PAD_MD)

    def _select(self, key):
        for btn in self._menu_btns.values():
            btn.configure(fg_color="transparent", text_color=TEXT_SECONDARY)
        # Solid dim highlight — no alpha
        self._menu_btns[key].configure(
            fg_color=NEON_BLUE_DIM,    # was NEON_BLUE + "22"
            text_color=NEON_BLUE
        )
        self._active = key
        self._load_page(key)

    # ──────────────────────────────────────────────────────────────────────────
    #  TOP BAR
    # ──────────────────────────────────────────────────────────────────────────
    def _build_topbar(self):
        bar = ctk.CTkFrame(self, height=60,
                            fg_color=BG_SURFACE, corner_radius=0)
        bar.pack(fill="x")

        self._page_title = ctk.CTkLabel(
            bar, text="Dashboard",
            font=FONT_HEADING, text_color=TEXT_PRIMARY
        )
        self._page_title.pack(side="left", padx=PAD_LG, pady=PAD_MD)

        self._clock_label = ctk.CTkLabel(
            bar, text="", font=("Courier New", 13), text_color=NEON_BLUE
        )
        self._clock_label.pack(side="right", padx=PAD_LG)
        self._tick_clock()

    def _tick_clock(self):
        self._clock_label.configure(
            text=datetime.now().strftime("%Y-%m-%d  %H:%M:%S")
        )
        self.after(1000, self._tick_clock)

    # ──────────────────────────────────────────────────────────────────────────
    #  BODY
    # ──────────────────────────────────────────────────────────────────────────
    def _build_body(self):
        self._body = ctk.CTkScrollableFrame(
            self, fg_color=BG_BASE, corner_radius=0
        )
        self._body.pack(fill="both", expand=True, padx=PAD_LG, pady=PAD_LG)

    def _clear_body(self):
        for w in self._body.winfo_children():
            w.destroy()

    def _load_page(self, key):
        self._clear_body()
        titles = {
            "dashboard":  "Dashboard",
            "employees":  "Employee Management",
            "attendance": "Attendance Records",
            "analytics":  "Analytics",
            "logs":       "System Logs",
            "settings":   "Settings",
        }
        self._page_title.configure(text=titles[key])
        {
            "dashboard":  self._page_dashboard,
            "employees":  self._page_employees,
            "attendance": self._page_attendance,
            "analytics":  self._page_analytics,
            "logs":       self._page_logs,
            "settings":   self._page_settings,
        }[key]()

    # ──────────────────────────────────────────────────────────────────────────
    #  DASHBOARD
    # ──────────────────────────────────────────────────────────────────────────
    def _page_dashboard(self):
        today   = datetime.now().strftime("%Y-%m-%d")
        total   = employees.count_documents({})
        present = attendance.count_documents({"date": today})
        late    = attendance.count_documents({
            "date": today, "login_time": {"$gt": LATE_AFTER}
        })
        absent  = max(total - present, 0)
        rate    = round(present / total * 100, 1) if total else 0

        # 5 stat cards
        row = ctk.CTkFrame(self._body, fg_color="transparent")
        row.pack(fill="x", pady=(0, PAD_LG))
        row.columnconfigure((0, 1, 2, 3, 4), weight=1)

        for i, (title, val, icon, accent) in enumerate([
            ("Total Employees", total,       "👥", NEON_BLUE),
            ("Present Today",   present,     "✅", NEON_GREEN),
            ("Late Arrivals",   late,        "⏰", NEON_YELLOW),
            ("Absent Today",    absent,      "❌", NEON_PINK),
            ("Attendance Rate", f"{rate}%",  "📈", NEON_PURPLE),
        ]):
            StatCard(row, title, val, icon, accent=accent,
                     width=200, height=155).grid(
                row=0, column=i, padx=PAD_XS, sticky="nsew"
            )

        # Attendance rate progress bar
        rate_card = ctk.CTkFrame(self._body, fg_color=BG_SURFACE,
                                  corner_radius=RADIUS_LG)
        rate_card.pack(fill="x", pady=(0, PAD_LG))

        ctk.CTkLabel(rate_card, text=f"Attendance Rate Today — {rate}%",
                     font=FONT_SUBHEAD, text_color=NEON_PURPLE).pack(
            anchor="w", padx=PAD_MD, pady=(PAD_MD, PAD_SM))

        bar_bg = ctk.CTkFrame(rate_card, height=18,
                               fg_color=BG_ELEVATED, corner_radius=9)
        bar_bg.pack(fill="x", padx=PAD_MD, pady=(0, PAD_MD))
        bar_bg.update_idletasks()
        bar_w = max(int(bar_bg.winfo_reqwidth() * rate / 100), 4)
        rate_color = (NEON_GREEN  if rate >= 80
                      else NEON_YELLOW if rate >= 50
                      else NEON_PINK)
        ctk.CTkFrame(bar_bg, width=bar_w, height=18,
                      fg_color=rate_color, corner_radius=9).pack(side="left")

        # Recent activity
        ctk.CTkLabel(self._body, text="Recent Activity",
                     font=FONT_HEADING,
                     text_color=TEXT_PRIMARY).pack(anchor="w", pady=(0, PAD_SM))
        container = ctk.CTkFrame(self._body, fg_color=BG_SURFACE,
                                  corner_radius=RADIUS_LG)
        container.pack(fill="both", expand=True)
        self._render_attendance_table(container, limit=15)

    # ──────────────────────────────────────────────────────────────────────────
    #  EMPLOYEES
    # ──────────────────────────────────────────────────────────────────────────
    def _page_employees(self):
        hrow = ctk.CTkFrame(self._body, fg_color="transparent")
        hrow.pack(fill="x", pady=(0, PAD_MD))
        ctk.CTkLabel(hrow, text="All Employees",
                     font=FONT_HEADING, text_color=TEXT_PRIMARY).pack(side="left")
        NeonButton(hrow, text="+ Add Employee",
                   accent=NEON_GREEN, width=160, height=40,
                   command=self._add_employee_dialog).pack(side="right")

        search_bar = ctk.CTkFrame(self._body, fg_color=BG_SURFACE,
                                   corner_radius=RADIUS_MD)
        search_bar.pack(fill="x", pady=(0, PAD_MD))
        ctk.CTkLabel(search_bar, text="🔍", font=FONT_BODY).pack(
            side="left", padx=PAD_SM)
        self._emp_search = ctk.CTkEntry(
            search_bar, placeholder_text="Search by ID or name…",
            fg_color="transparent", border_width=0,
            font=FONT_BODY, width=300)
        self._emp_search.pack(side="left", pady=PAD_SM)
        NeonButton(search_bar, text="Search",
                   accent=NEON_BLUE, width=80, height=32,
                   command=self._search_employees).pack(side="left", padx=PAD_SM)

        self._emp_table_frame = ctk.CTkFrame(
            self._body, fg_color=BG_SURFACE, corner_radius=RADIUS_LG)
        self._emp_table_frame.pack(fill="both", expand=True)
        self._render_employee_table()

    def _render_employee_table(self, query=None):
        for w in self._emp_table_frame.winfo_children():
            w.destroy()

        apply_tree_style()
        cols = ("ID", "Name", "Photo", "Action")
        tree = ttk.Treeview(self._emp_table_frame, columns=cols,
                            show="headings", style="Cyber.Treeview", height=15)
        for col, w in [("ID", 130), ("Name", 240), ("Photo", 100), ("Action", 160)]:
            tree.heading(col, text=col)
            tree.column(col, width=w)

        sb = ctk.CTkScrollbar(self._emp_table_frame, command=tree.yview)
        sb.pack(side="right", fill="y")
        tree.configure(yscrollcommand=sb.set)
        tree.pack(fill="both", expand=True, padx=PAD_SM, pady=PAD_SM)

        for emp in employees.find(query or {}).sort("emp_id", 1):
            has_photo = "📸 Yes" if os.path.exists(
                os.path.join(FACES_DIR, f"{emp['emp_id']}.jpg")) else "—"
            tree.insert("", "end", values=(
                emp["emp_id"], emp["emp_name"], has_photo, "Double-click →"
            ))

        tree.bind("<Double-1>", lambda e: self._emp_row_action(tree))

    def _emp_row_action(self, tree):
        sel = tree.selection()
        if not sel: return
        emp_id, emp_name = tree.item(sel[0], "values")[:2]
        self._action_popup(emp_id, emp_name)

    def _search_employees(self):
        q = self._emp_search.get().strip()
        query = {"$or": [
            {"emp_id":   {"$regex": q, "$options": "i"}},
            {"emp_name": {"$regex": q, "$options": "i"}},
        ]} if q else None
        self._render_employee_table(query)

    # ── Action popup  ─────────────────────────────────────────────────────────
    def _action_popup(self, emp_id, emp_name):
        win = ctk.CTkToplevel(self)
        win.title("Employee Actions")
        win.geometry("360x420")
        win.transient(self); win.grab_set()
        win.configure(fg_color=BG_ELEVATED)

        # Photo preview
        photo_path = os.path.join(FACES_DIR, f"{emp_id}.jpg")
        if os.path.exists(photo_path):
            try:
                from PIL import Image
                img     = Image.open(photo_path).resize((100, 100))
                ctk_img = ctk.CTkImage(img, size=(100, 100))
                ctk.CTkLabel(win, image=ctk_img, text="",
                             corner_radius=RADIUS_MD).pack(pady=(PAD_MD, PAD_SM))
            except Exception:
                ctk.CTkLabel(win, text="👤", font=("Segoe UI", 56)).pack(
                    pady=(PAD_MD, PAD_SM))
        else:
            ctk.CTkLabel(win, text="👤", font=("Segoe UI", 56)).pack(
                pady=(PAD_MD, PAD_SM))

        ctk.CTkLabel(win, text=emp_name, font=FONT_HEADING,
                     text_color=NEON_BLUE).pack()
        ctk.CTkLabel(win, text=f"ID: {emp_id}", font=FONT_SMALL,
                     text_color=TEXT_SECONDARY).pack(pady=(2, PAD_SM))

        NeonButton(win, text="✏️  Edit Details",
                   accent=NEON_YELLOW, height=40,
                   command=lambda: [win.destroy(),
                                    self._edit_employee(emp_id, emp_name)]
                   ).pack(fill="x", padx=PAD_LG, pady=4)

        NeonButton(win, text="📷  Update Face",
                   accent=NEON_BLUE, height=40,
                   command=lambda: [win.destroy(),
                                    self._update_face(emp_id, emp_name)]
                   ).pack(fill="x", padx=PAD_LG, pady=4)

        # ── DELETE: confirm FIRST, destroy window AFTER ────────────────────
        def _confirm_delete():
            # Ask confirmation while popup is still visible
            confirmed = messagebox.askyesno(
                "Confirm Delete",
                f"Delete '{emp_name}' and ALL their attendance records?\n\nThis cannot be undone.",
                parent=win
            )
            if confirmed:
                win.destroy()                          # close popup after confirm
                employees.delete_one({"emp_id": emp_id})
                attendance.delete_many({"emp_id": emp_id})
                photo = os.path.join(FACES_DIR, f"{emp_id}.jpg")
                if os.path.exists(photo):
                    os.remove(photo)
                write_log("Employee Deleted", f"{emp_name} ({emp_id})", "WARNING")
                self._render_employee_table()          # refresh table
                messagebox.showinfo("Deleted", f"'{emp_name}' has been removed.")

        NeonButton(win, text="🗑️  Delete Employee",
                   accent=NEON_PINK, height=40,
                   command=_confirm_delete
                   ).pack(fill="x", padx=PAD_LG, pady=4)

        ctk.CTkButton(win, text="Cancel",
                      fg_color=BG_SURFACE, hover_color=BG_HOVER,
                      command=win.destroy, height=36).pack(pady=PAD_SM)

    # ── Edit / Update Face ────────────────────────────────────────────────────
    def _edit_employee(self, emp_id, emp_name):
        win = ctk.CTkToplevel(self)
        win.title("Edit Employee")
        win.geometry("380x280")
        win.transient(self); win.grab_set()
        win.configure(fg_color=BG_ELEVATED)

        ctk.CTkLabel(win, text="Edit Employee",
                     font=FONT_HEADING, text_color=NEON_BLUE).pack(pady=PAD_MD)

        entries = {}
        for label, default in [("Employee ID", emp_id), ("Employee Name", emp_name)]:
            ctk.CTkLabel(win, text=label, font=FONT_SMALL,
                         text_color=TEXT_SECONDARY).pack(anchor="w", padx=PAD_LG)
            e = ctk.CTkEntry(win, width=300)
            e.insert(0, default)
            e.pack(padx=PAD_LG, pady=(0, PAD_SM))
            entries[label] = e

        def save():
            nid   = entries["Employee ID"].get().strip()
            nname = entries["Employee Name"].get().strip()
            if not nid or not nname:
                messagebox.showwarning("Warning", "All fields required"); return
            employees.update_one({"emp_id": emp_id},
                                  {"$set": {"emp_id": nid, "emp_name": nname}})
            attendance.update_many({"emp_id": emp_id},
                                    {"$set": {"emp_id": nid, "emp_name": nname}})
            write_log("Employee Edit", f"{emp_name} → {nname} ({nid})", "INFO")
            win.destroy()
            self._page_employees()
            messagebox.showinfo("Saved", "Employee updated ✔")

        NeonButton(win, text="Save Changes",
                   accent=NEON_GREEN, command=save, height=40).pack(
            padx=PAD_LG, fill="x", pady=PAD_SM)

    def _update_face(self, emp_id, emp_name):
        from register import capture_face_encoding
        enc = capture_face_encoding(emp_name, emp_id)
        if enc is None:
            messagebox.showerror("Error", "Face capture failed"); return
        employees.update_one({"emp_id": emp_id},
                              {"$set": {"face_encoding": enc.tolist()}})
        write_log("Face Update", f"{emp_name} face recaptured", "INFO")
        messagebox.showinfo("Updated", f"Face updated for {emp_name} ✔")

    def _add_employee_dialog(self):
        from register import capture_face_encoding
        win = ctk.CTkToplevel(self)
        win.title("Add Employee")
        win.geometry("400x340")
        win.transient(self); win.grab_set()
        win.configure(fg_color=BG_ELEVATED)

        ctk.CTkLabel(win, text="Add New Employee",
                     font=FONT_HEADING, text_color=NEON_GREEN).pack(pady=PAD_MD)

        entries = {}
        for label in ["Employee ID", "Employee Name"]:
            ctk.CTkLabel(win, text=label, font=FONT_SMALL,
                         text_color=TEXT_SECONDARY).pack(anchor="w", padx=PAD_LG)
            e = ctk.CTkEntry(win, width=300)
            e.pack(padx=PAD_LG, pady=(0, PAD_SM))
            entries[label] = e

        def add():
            eid   = entries["Employee ID"].get().strip()
            ename = entries["Employee Name"].get().strip()
            if not eid or not ename:
                messagebox.showwarning("Warning", "All fields required"); return
            if employees.find_one({"emp_id": eid}):
                messagebox.showerror("Error", "ID already exists"); return
            enc = capture_face_encoding(ename, eid)
            if enc is None:
                messagebox.showerror("Error", "Face capture failed"); return
            employees.insert_one({
                "emp_id": eid, "emp_name": ename,
                "face_encoding": enc.tolist()
            })
            write_log("Employee Added", f"{ename} ({eid})", "SUCCESS")
            win.destroy()
            self._page_employees()
            messagebox.showinfo("Success", f"{ename} added ✔")

        NeonButton(win, text="📷  Capture Face & Save",
                   accent=NEON_GREEN, command=add, height=46).pack(
            padx=PAD_LG, fill="x", pady=PAD_MD)

    # ──────────────────────────────────────────────────────────────────────────
    #  ATTENDANCE
    # ──────────────────────────────────────────────────────────────────────────
    def _page_attendance(self):
        frow = ctk.CTkFrame(self._body, fg_color=BG_SURFACE, corner_radius=RADIUS_MD)
        frow.pack(fill="x", pady=(0, PAD_MD))

        ctk.CTkLabel(frow, text="Filter date:", font=FONT_SMALL,
                     text_color=TEXT_SECONDARY).pack(side="left", padx=PAD_SM)
        date_entry = ctk.CTkEntry(frow, width=140, placeholder_text="YYYY-MM-DD")
        date_entry.pack(side="left", padx=PAD_SM, pady=PAD_SM)

        table_container = ctk.CTkFrame(self._body, fg_color=BG_SURFACE,
                                        corner_radius=RADIUS_LG)

        NeonButton(frow, text="Apply",
                   accent=NEON_BLUE, width=80, height=34,
                   command=lambda: self._render_attendance_table(
                       table_container,
                       filter_date=date_entry.get().strip() or None
                   )).pack(side="left")
        NeonButton(frow, text="Clear",
                   accent=NEON_PURPLE, width=70, height=34,
                   command=lambda: [date_entry.delete(0, "end"),
                                    self._render_attendance_table(table_container)]
                   ).pack(side="left", padx=PAD_SM)
        NeonButton(frow, text="⬇ Excel",
                   accent=NEON_GREEN, width=100, height=34,
                   command=self._export_excel).pack(side="right", padx=PAD_SM)
        NeonButton(frow, text="⬇ CSV",
                   accent=NEON_BLUE, width=90, height=34,
                   command=self._export_csv).pack(side="right", padx=PAD_SM)

        table_container.pack(fill="both", expand=True)
        self._render_attendance_table(table_container)

    def _render_attendance_table(self, parent, filter_date=None, limit=None):
        for w in parent.winfo_children():
            w.destroy()

        apply_tree_style()
        cols = ("ID", "Name", "Date", "Login", "Logout", "Hours", "Late", "Status")
        tree = ttk.Treeview(parent, columns=cols,
                            show="headings", style="Cyber.Treeview",
                            height=16 if limit is None else limit)
        for col, w in zip(cols, (75, 150, 95, 80, 80, 70, 80, 90)):
            tree.heading(col, text=col)
            tree.column(col, width=w)

        sb = ctk.CTkScrollbar(parent, command=tree.yview)
        sb.pack(side="right", fill="y")
        tree.configure(yscrollcommand=sb.set)
        tree.pack(fill="both", expand=True, padx=PAD_SM, pady=PAD_SM)

        query  = {"date": filter_date} if filter_date else {}
        cursor = attendance.find(query).sort([("date", -1), ("login_time", -1)])
        if limit:
            cursor = cursor.limit(limit)

        for rec in cursor:
            hours_str, late_str, status = self._calc_row(rec)
            tree.insert("", "end", values=(
                rec["emp_id"], rec["emp_name"], rec["date"],
                rec.get("login_time", "—"), rec.get("logout_time", "—"),
                hours_str, late_str, status
            ))

    # ──────────────────────────────────────────────────────────────────────────
    #  HELPERS
    # ──────────────────────────────────────────────────────────────────────────
    def _calc_row(self, rec):
        login_str  = rec.get("login_time",  "")
        logout_str = rec.get("logout_time", "")

        # Late detection
        late_str = "—"
        if login_str:
            try:
                threshold = datetime.strptime(LATE_AFTER, "%H:%M:%S")
                login_dt  = datetime.strptime(login_str,  "%H:%M:%S")
                late_str  = "🔴 Late" if login_dt > threshold else "🟢 On Time"
            except Exception:
                late_str = "—"

        # Hours worked
        hours_str = "—"
        status    = "Active"
        if login_str and logout_str:
            try:
                login_dt  = datetime.strptime(login_str,  "%H:%M:%S")
                logout_dt = datetime.strptime(logout_str, "%H:%M:%S")
                hours     = round((logout_dt - login_dt).total_seconds() / 3600, 2)
                hours_str = f"{hours} h"
                status = ("Completed" if hours >= 8
                          else "Partial"   if hours >= 4
                          else "Short")
            except Exception:
                hours_str, status = "—", "Unknown"
        elif login_str and not logout_str:
            status = "Active"

        return hours_str, late_str, status

    # ──────────────────────────────────────────────────────────────────────────
    #  EXPORT
    # ──────────────────────────────────────────────────────────────────────────
    def _export_csv(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".csv", filetypes=[("CSV", "*.csv")])
        if not path: return
        with open(path, "w", newline="") as f:
            wr = csv.writer(f)
            wr.writerow(["ID", "Name", "Date", "Login",
                         "Logout", "Hours", "Late", "Status"])
            for rec in attendance.find().sort([("date", -1)]):
                h, l, s = self._calc_row(rec)
                wr.writerow([rec["emp_id"], rec["emp_name"], rec["date"],
                             rec.get("login_time", ""), rec.get("logout_time", ""),
                             h, l.replace("🔴 ", "").replace("🟢 ", ""), s])
        write_log("Export", f"CSV → {path}", "INFO")
        messagebox.showinfo("Export", f"CSV saved ✔\n{path}")

    def _export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            messagebox.showerror("Missing",
                                 "Run:  pip install openpyxl"); return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")])
        if not path: return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance"

        headers = ["ID", "Name", "Date", "Login",
                   "Logout", "Hours Worked", "Late/On Time", "Status"]
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        hdr_fill = PatternFill("solid", fgColor="0D1B2A")
        thin     = Side(border_style="thin", color="333333")
        border   = Border(left=thin, right=thin, top=thin, bottom=thin)

        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border    = border
        ws.row_dimensions[1].height = 22

        status_fills = {
            "Completed": PatternFill("solid", fgColor="0D3B1E"),
            "Partial":   PatternFill("solid", fgColor="1E3B4A"),
            "Short":     PatternFill("solid", fgColor="3B1E1E"),
            "Active":    PatternFill("solid", fgColor="0D1E3B"),
        }

        for ri, rec in enumerate(
            attendance.find().sort([("date", -1)]), start=2
        ):
            h, l, s = self._calc_row(rec)
            l_clean  = l.replace("🔴 ", "").replace("🟢 ", "")
            row_data = [rec["emp_id"], rec["emp_name"], rec["date"],
                        rec.get("login_time", ""), rec.get("logout_time", ""),
                        h, l_clean, s]
            for ci, val in enumerate(row_data, 1):
                cell           = ws.cell(row=ri, column=ci, value=val)
                cell.border    = border
                cell.alignment = Alignment(horizontal="center")
                if ci == 8 and s in status_fills:
                    cell.fill  = status_fills[s]
            ws.row_dimensions[ri].height = 18

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

        wb.save(path)
        write_log("Export", f"Excel → {path}", "INFO")
        messagebox.showinfo("Export", f"Excel saved ✔\n{path}")

    # ──────────────────────────────────────────────────────────────────────────
    #  ANALYTICS — 30-day trend + pie + bar
    # ──────────────────────────────────────────────────────────────────────────
    def _page_analytics(self):
        try:
            from matplotlib.figure import Figure
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            import matplotlib
        except ImportError:
            ctk.CTkLabel(self._body, text="pip install matplotlib",
                         text_color=NEON_PINK).pack(); return

        today   = datetime.now().strftime("%Y-%m-%d")
        total   = employees.count_documents({})
        present = attendance.count_documents({"date": today})
        absent  = max(total - present, 0)
        late    = attendance.count_documents({
            "date": today, "login_time": {"$gt": LATE_AFTER}})
        rate    = round(present / total * 100, 1) if total else 0

        # Stat cards
        row = ctk.CTkFrame(self._body, fg_color="transparent")
        row.pack(fill="x", pady=(0, PAD_LG))
        row.columnconfigure((0, 1, 2, 3), weight=1)
        for i, (title, val, accent) in enumerate([
            ("Present Today",    present,    NEON_GREEN),
            ("Absent Today",     absent,     NEON_PINK),
            ("Late Today",       late,       NEON_YELLOW),
            ("Attendance Rate",  f"{rate}%", NEON_PURPLE),
        ]):
            StatCard(row, title, val, accent=accent).grid(
                row=0, column=i, padx=PAD_SM, sticky="nsew")

        # Chart bg (must be plain string for matplotlib)
        mode   = ctk.get_appearance_mode()
        plt_bg = BG_SURFACE_D if mode == "Dark" else BG_SURFACE_L

        # Row 1: pie + 7-day bar side by side
        charts1 = ctk.CTkFrame(self._body, fg_color="transparent")
        charts1.pack(fill="both", pady=(0, PAD_MD))

        # Pie
        pie_f = ctk.CTkFrame(charts1, fg_color=BG_SURFACE, corner_radius=RADIUS_LG)
        pie_f.pack(side="left", fill="both", expand=True, padx=(0, PAD_SM))
        ctk.CTkLabel(pie_f, text="Today's Breakdown",
                     font=FONT_SUBHEAD, text_color=NEON_BLUE).pack(pady=(PAD_MD, 0))
        fig_pie = Figure(figsize=(3.5, 3.2), dpi=90, facecolor=plt_bg)
        ax_pie  = fig_pie.add_subplot(111, facecolor=plt_bg)
        ax_pie.pie([present or 1, absent or 1, late or 0.001],
                   labels=["Present", "Absent", "Late"],
                   colors=[NEON_GREEN, NEON_PINK, NEON_YELLOW],
                   autopct="%1.0f%%", startangle=90,
                   textprops={"color": "white", "fontsize": 9})
        FigureCanvasTkAgg(fig_pie, pie_f).get_tk_widget().pack(
            fill="both", expand=True, padx=PAD_SM, pady=PAD_SM)

        # 7-day bar
        bar_f = ctk.CTkFrame(charts1, fg_color=BG_SURFACE, corner_radius=RADIUS_LG)
        bar_f.pack(side="right", fill="both", expand=True)
        ctk.CTkLabel(bar_f, text="Last 7 Days",
                     font=FONT_SUBHEAD, text_color=NEON_BLUE).pack(pady=(PAD_MD, 0))
        days7, vals7 = [], []
        for i in range(6, -1, -1):
            d = datetime.now() - timedelta(days=i)
            days7.append(d.strftime("%a"))
            vals7.append(attendance.count_documents({"date": d.strftime("%Y-%m-%d")}))

        fig_bar = Figure(figsize=(4.5, 3.2), dpi=90, facecolor=plt_bg)
        ax_bar  = fig_bar.add_subplot(111, facecolor=plt_bg)
        ax_bar.bar(days7, vals7,
                   color=[NEON_BLUE, NEON_PURPLE, NEON_GREEN,
                          NEON_PINK, NEON_ORANGE, NEON_BLUE, NEON_PURPLE],
                   width=0.6)
        ax_bar.tick_params(colors="white", labelsize=9)
        for sp in ax_bar.spines.values(): sp.set_color(BG_ELEVATED_D)
        ax_bar.yaxis.set_major_locator(
            matplotlib.ticker.MaxNLocator(integer=True))
        FigureCanvasTkAgg(fig_bar, bar_f).get_tk_widget().pack(
            fill="both", expand=True, padx=PAD_SM, pady=PAD_SM)

        # Row 2: 30-day trend line
        trend_f = ctk.CTkFrame(self._body, fg_color=BG_SURFACE, corner_radius=RADIUS_LG)
        trend_f.pack(fill="both", expand=True)
        ctk.CTkLabel(trend_f, text="30-Day Attendance Trend",
                     font=FONT_SUBHEAD, text_color=NEON_PURPLE).pack(
            anchor="w", padx=PAD_MD, pady=(PAD_MD, 0))

        days30, vals30 = [], []
        for i in range(29, -1, -1):
            d = datetime.now() - timedelta(days=i)
            days30.append(d.strftime("%m/%d"))
            vals30.append(attendance.count_documents({"date": d.strftime("%Y-%m-%d")}))

        fig30 = Figure(figsize=(9, 2.8), dpi=90, facecolor=plt_bg)
        ax30  = fig30.add_subplot(111, facecolor=plt_bg)
        ax30.plot(days30, vals30, color=NEON_PURPLE, linewidth=2,
                  marker="o", markersize=4, markerfacecolor=NEON_BLUE)
        ax30.fill_between(range(len(days30)), vals30, alpha=0.15, color=NEON_PURPLE)
        ax30.set_title("Daily Headcount — Last 30 Days",
                       color="white", fontsize=10, pad=6)
        ax30.tick_params(colors="white", labelsize=7)
        ax30.set_xticks(range(0, len(days30), 3))
        ax30.set_xticklabels(days30[::3], rotation=35, ha="right")
        for sp in ax30.spines.values(): sp.set_color(BG_ELEVATED_D)
        ax30.yaxis.set_major_locator(matplotlib.ticker.MaxNLocator(integer=True))
        fig30.tight_layout()
        FigureCanvasTkAgg(fig30, trend_f).get_tk_widget().pack(
            fill="both", expand=True, padx=PAD_SM, pady=PAD_SM)

    # ──────────────────────────────────────────────────────────────────────────
    #  SYSTEM LOGS
    # ──────────────────────────────────────────────────────────────────────────
    def _page_logs(self):
        hrow = ctk.CTkFrame(self._body, fg_color="transparent")
        hrow.pack(fill="x", pady=(0, PAD_MD))
        ctk.CTkLabel(hrow, text="System Event Logs",
                     font=FONT_HEADING, text_color=TEXT_PRIMARY).pack(side="left")
        NeonButton(hrow, text="🗑 Clear All Logs",
                   accent=NEON_PINK, width=160, height=38,
                   command=self._clear_logs).pack(side="right")

        frow = ctk.CTkFrame(self._body, fg_color=BG_SURFACE, corner_radius=RADIUS_MD)
        frow.pack(fill="x", pady=(0, PAD_MD))
        ctk.CTkLabel(frow, text="Level:", font=FONT_SMALL,
                     text_color=TEXT_SECONDARY).pack(side="left", padx=PAD_SM)

        self._log_filter = ctk.CTkOptionMenu(
            frow,
            values=["ALL", "SUCCESS", "INFO", "WARNING", "ERROR"],
            command=lambda _: self._reload_logs(log_container)
        )
        self._log_filter.set("ALL")
        self._log_filter.pack(side="left", padx=PAD_SM, pady=PAD_SM)

        log_container = ctk.CTkFrame(self._body, fg_color=BG_SURFACE,
                                      corner_radius=RADIUS_LG)
        log_container.pack(fill="both", expand=True)
        self._reload_logs(log_container)

    def _reload_logs(self, parent):
        for w in parent.winfo_children():
            w.destroy()

        apply_tree_style()
        cols = ("Timestamp", "Level", "Event", "Detail")
        tree = ttk.Treeview(parent, columns=cols,
                            show="headings", style="Cyber.Treeview", height=20)
        for col, w in [("Timestamp", 160), ("Level", 90),
                        ("Event", 160), ("Detail", 380)]:
            tree.heading(col, text=col)
            tree.column(col, width=w)

        sb = ctk.CTkScrollbar(parent, command=tree.yview)
        sb.pack(side="right", fill="y")
        tree.configure(yscrollcommand=sb.set)
        tree.pack(fill="both", expand=True, padx=PAD_SM, pady=PAD_SM)

        level_filter = self._log_filter.get()
        query = {} if level_filter == "ALL" else {"level": level_filter}
        icons = {"SUCCESS": "✅", "INFO": "ℹ️", "WARNING": "⚠️", "ERROR": "❌"}

        for log in (logs or []).find(query).sort("timestamp", -1).limit(200):
            icon = icons.get(log.get("level", ""), "•")
            tree.insert("", "end", values=(
                log.get("timestamp", ""),
                f"{icon} {log.get('level', '')}",
                log.get("event", ""),
                log.get("detail", ""),
            ))

    def _clear_logs(self):
        if messagebox.askyesno("Confirm", "Delete all system logs?"):
            if logs: logs.delete_many({})
            self._page_logs()

    # ──────────────────────────────────────────────────────────────────────────
    #  SETTINGS — theme switch rebuilds BG correctly
    # ──────────────────────────────────────────────────────────────────────────
    def _page_settings(self):
        card = ctk.CTkFrame(self._body, fg_color=BG_SURFACE, corner_radius=RADIUS_LG)
        card.pack(fill="x", padx=PAD_SM, pady=PAD_SM)

        # ── Appearance ────────────────────────────────────────────────────────
        ctk.CTkLabel(card, text="Appearance", font=FONT_SUBHEAD,
                     text_color=NEON_BLUE).pack(
            anchor="w", padx=PAD_MD, pady=(PAD_MD, PAD_SM))
        ctk.CTkFrame(card, height=1,
                     fg_color=NEON_BLUE_MID).pack(fill="x", padx=PAD_MD)

        row = ctk.CTkFrame(card, fg_color="transparent")
        row.pack(fill="x", padx=PAD_MD, pady=PAD_SM)
        ctk.CTkLabel(row, text="Theme:", width=160,
                     text_color=TEXT_SECONDARY).pack(side="left")

        def apply_theme(mode):
            """Switch theme — CTk handles panel recolor via tuples automatically."""
            from ui.theme import set_theme
            set_theme(mode)
            # Refresh treeview colors (ttk doesn't auto-update)
            apply_tree_style()

        opt = ctk.CTkOptionMenu(row, values=["Dark", "Light", "System"],
                                 command=apply_theme)
        opt.set("Dark")
        opt.pack(side="left")

        # ── Face Recognition ──────────────────────────────────────────────────
        ctk.CTkLabel(card, text="Face Recognition", font=FONT_SUBHEAD,
                     text_color=NEON_PURPLE).pack(
            anchor="w", padx=PAD_MD, pady=(PAD_MD, PAD_SM))
        ctk.CTkFrame(card, height=1,
                     fg_color=NEON_PURPLE_MID).pack(fill="x", padx=PAD_MD)

        row2 = ctk.CTkFrame(card, fg_color="transparent")
        row2.pack(fill="x", padx=PAD_MD, pady=PAD_SM)
        ctk.CTkLabel(row2, text="Match Tolerance:", width=160,
                     text_color=TEXT_SECONDARY).pack(side="left")
        sl = ctk.CTkSlider(row2, from_=0.3, to=0.7, number_of_steps=40)
        sl.set(0.5)
        sl.pack(side="left", padx=PAD_SM)
        val_lbl = ctk.CTkLabel(row2, text="0.50", font=FONT_SMALL,
                               text_color=NEON_BLUE)
        val_lbl.pack(side="left")
        sl.configure(command=lambda v: val_lbl.configure(text=f"{float(v):.2f}"))

        # ── Late Threshold ────────────────────────────────────────────────────
        ctk.CTkLabel(card, text="Late Threshold", font=FONT_SUBHEAD,
                     text_color=NEON_ORANGE).pack(
            anchor="w", padx=PAD_MD, pady=(PAD_MD, PAD_SM))
        ctk.CTkFrame(card, height=1,
                     fg_color=NEON_ORANGE_DIM).pack(fill="x", padx=PAD_MD)

        row3 = ctk.CTkFrame(card, fg_color="transparent")
        row3.pack(fill="x", padx=PAD_MD, pady=PAD_SM)
        ctk.CTkLabel(row3, text="Late after:", width=160,
                     text_color=TEXT_SECONDARY).pack(side="left")
        late_entry = ctk.CTkEntry(row3, width=100)
        late_entry.insert(0, LATE_AFTER)
        late_entry.pack(side="left", padx=PAD_SM)
        ctk.CTkLabel(row3, text="(HH:MM:SS)",
                     font=FONT_MICRO, text_color=TEXT_MUTED).pack(side="left")

        NeonButton(card, text="Save Settings",
                   accent=NEON_GREEN,
                   command=lambda: messagebox.showinfo("Saved", "Settings saved ✔"),
                   height=42).pack(padx=PAD_MD, pady=PAD_MD)